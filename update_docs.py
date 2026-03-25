#!/usr/bin/env python3
"""Update table definition and API spec with complete data from agent analysis"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

DOCS = os.path.join(os.path.dirname(os.path.abspath(__file__)), "docs")
HEADER_FILL = PatternFill(start_color="1A478A", end_color="1A478A", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10, name="Malgun Gothic")
CELL_FONT = Font(size=10, name="Malgun Gothic")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

def write_sheet(ws, headers, rows, title=None):
    if title:
        ws.title = title
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = CELL_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
    for ci in range(1, len(headers) + 1):
        letter = get_column_letter(ci)
        max_len = max(len(str(ws.cell(row=r, column=ci).value or "")) for r in range(1, len(rows) + 2))
        ws.column_dimensions[letter].width = min(max_len + 4, 55)


def update_table_definition():
    wb = Workbook()
    ws = wb.active
    headers = ["테이블명", "컬럼명", "데이터타입", "PK", "FK", "NOT NULL", "DEFAULT", "설명"]
    data = [
        # users
        ["users", "id", "BIGSERIAL", "Y", "", "Y", "", "사용자 ID"],
        ["users", "email", "VARCHAR(255)", "", "", "Y", "", "이메일 (UNIQUE)"],
        ["users", "password", "VARCHAR(255)", "", "", "Y", "", "BCrypt 해싱 비밀번호"],
        ["users", "role", "VARCHAR(50)", "", "", "Y", "USER", "역할 (USER/ADMIN)"],
        ["users", "status", "VARCHAR(20)", "", "", "Y", "ACTIVE", "사용자 상태 (V8)"],
        ["users", "created_at", "TIMESTAMP", "", "", "Y", "NOW()", "생성일시"],
        ["users", "updated_at", "TIMESTAMP", "", "", "Y", "NOW()", "수정일시"],
        # orders
        ["orders", "id", "BIGSERIAL", "Y", "", "Y", "", "주문 ID"],
        ["orders", "user_id", "BIGINT", "", "users(id)", "Y", "", "주문자"],
        ["orders", "product_id", "BIGINT", "", "products(id)", "", "", "상품 (V15)"],
        ["orders", "amount", "DECIMAL(10,2)", "", "", "Y", "", "주문 금액"],
        ["orders", "status", "VARCHAR(20)", "", "", "Y", "CREATED", "주문상태"],
        ["orders", "created_at", "TIMESTAMP", "", "", "Y", "NOW()", "주문일시"],
        ["orders", "updated_at", "TIMESTAMP", "", "", "Y", "NOW()", "수정일시"],
        # payments
        ["payments", "id", "BIGSERIAL", "Y", "", "Y", "", "결제 ID"],
        ["payments", "order_id", "BIGINT", "", "orders(id)", "Y", "", "주문 (UNIQUE active)"],
        ["payments", "amount", "DECIMAL(10,2)", "", "", "Y", "", "결제금액"],
        ["payments", "refunded_amount", "DECIMAL(10,2)", "", "", "Y", "0", "환불금액"],
        ["payments", "status", "VARCHAR(20)", "", "", "Y", "READY", "결제상태"],
        ["payments", "payment_method", "VARCHAR(50)", "", "", "", "", "결제수단"],
        ["payments", "pg_transaction_id", "VARCHAR(500)", "", "", "", "", "Toss paymentKey (V16)"],
        ["payments", "captured_at", "TIMESTAMP", "", "", "", "", "매입일시"],
        ["payments", "created_at", "TIMESTAMP", "", "", "Y", "NOW()", "생성일시"],
        ["payments", "updated_at", "TIMESTAMP", "", "", "Y", "NOW()", "수정일시"],
        # settlements
        ["settlements", "id", "BIGSERIAL", "Y", "", "Y", "", "정산 ID"],
        ["settlements", "payment_id", "BIGINT", "", "payments(id)", "Y", "", "결제 (UNIQUE)"],
        ["settlements", "order_id", "BIGINT", "", "orders(id)", "Y", "", "주문"],
        ["settlements", "payment_amount", "DECIMAL(10,2)", "", "", "Y", "", "결제금액"],
        ["settlements", "commission", "DECIMAL(10,2)", "", "", "Y", "", "수수료 (3%)"],
        ["settlements", "net_amount", "DECIMAL(10,2)", "", "", "Y", "", "순수익"],
        ["settlements", "status", "VARCHAR(20)", "", "", "Y", "PENDING", "정산상태"],
        ["settlements", "settlement_date", "DATE", "", "", "Y", "", "정산일"],
        ["settlements", "confirmed_at", "TIMESTAMP", "", "", "", "", "확정일시"],
        ["settlements", "approved_by", "BIGINT", "", "", "", "", "승인자 (V7)"],
        ["settlements", "approved_at", "TIMESTAMP", "", "", "", "", "승인일시 (V7)"],
        ["settlements", "rejected_by", "BIGINT", "", "", "", "", "거부자 (V7)"],
        ["settlements", "rejected_at", "TIMESTAMP", "", "", "", "", "거부일시 (V7)"],
        ["settlements", "rejection_reason", "VARCHAR(500)", "", "", "", "", "거부사유 (V7)"],
        ["settlements", "created_at", "TIMESTAMP", "", "", "Y", "NOW()", "생성일시"],
        ["settlements", "updated_at", "TIMESTAMP", "", "", "Y", "NOW()", "수정일시"],
        # refunds
        ["refunds", "id", "BIGSERIAL", "Y", "", "Y", "", "환불 ID"],
        ["refunds", "payment_id", "BIGINT", "", "payments(id)", "Y", "", "결제"],
        ["refunds", "amount", "DECIMAL(10,2)", "", "", "Y", "", "환불금액 (>0)"],
        ["refunds", "status", "VARCHAR(20)", "", "", "Y", "REQUESTED", "환불상태"],
        ["refunds", "reason", "TEXT", "", "", "", "", "환불사유"],
        ["refunds", "idempotency_key", "VARCHAR(255)", "", "", "Y", "", "멱등성 키"],
        ["refunds", "requested_at", "TIMESTAMP", "", "", "Y", "NOW()", "요청일시"],
        ["refunds", "completed_at", "TIMESTAMP", "", "", "", "", "완료일시"],
        ["refunds", "created_at", "TIMESTAMP", "", "", "Y", "NOW()", "생성일시"],
        ["refunds", "updated_at", "TIMESTAMP", "", "", "Y", "NOW()", "수정일시"],
        # settlement_adjustments
        ["settlement_adjustments", "id", "BIGSERIAL", "Y", "", "Y", "", "조정 ID"],
        ["settlement_adjustments", "settlement_id", "BIGINT", "", "settlements(id)", "Y", "", "정산"],
        ["settlement_adjustments", "refund_id", "BIGINT", "", "refunds(id)", "Y", "", "환불 (UNIQUE)"],
        ["settlement_adjustments", "amount", "DECIMAL(10,2)", "", "", "Y", "", "조정금액 (<0)"],
        ["settlement_adjustments", "status", "VARCHAR(20)", "", "", "Y", "PENDING", "조정상태"],
        ["settlement_adjustments", "adjustment_date", "DATE", "", "", "Y", "", "조정일"],
        ["settlement_adjustments", "confirmed_at", "TIMESTAMP", "", "", "", "", "확정일시"],
        ["settlement_adjustments", "created_at", "TIMESTAMP", "", "", "Y", "NOW()", "생성일시"],
        ["settlement_adjustments", "updated_at", "TIMESTAMP", "", "", "Y", "NOW()", "수정일시"],
        # batch_run_history
        ["batch_run_history", "id", "BIGSERIAL", "Y", "", "Y", "", "ID"],
        ["batch_run_history", "batch_name", "VARCHAR(100)", "", "", "Y", "", "배치명"],
        ["batch_run_history", "run_id", "VARCHAR(100)", "", "", "Y", "", "실행ID"],
        ["batch_run_history", "target_date", "DATE", "", "", "Y", "", "대상일"],
        ["batch_run_history", "status", "VARCHAR(20)", "", "", "Y", "", "상태"],
        ["batch_run_history", "started_at", "TIMESTAMP", "", "", "Y", "NOW()", "시작일시"],
        ["batch_run_history", "completed_at", "TIMESTAMP", "", "", "", "", "완료일시"],
        ["batch_run_history", "processed_count", "INT", "", "", "", "0", "처리건수"],
        ["batch_run_history", "error_message", "TEXT", "", "", "", "", "에러메시지"],
        # settlement_index_queue
        ["settlement_index_queue", "id", "BIGSERIAL", "Y", "", "Y", "", "ID"],
        ["settlement_index_queue", "settlement_id", "BIGINT", "", "", "Y", "", "정산ID"],
        ["settlement_index_queue", "operation", "VARCHAR(20)", "", "", "Y", "", "작업 (INDEX/UPDATE/DELETE)"],
        ["settlement_index_queue", "retry_count", "INTEGER", "", "", "Y", "0", "재시도횟수"],
        ["settlement_index_queue", "max_retries", "INTEGER", "", "", "Y", "3", "최대재시도"],
        ["settlement_index_queue", "status", "VARCHAR(20)", "", "", "Y", "PENDING", "상태"],
        ["settlement_index_queue", "error_message", "TEXT", "", "", "", "", "에러메시지"],
        ["settlement_index_queue", "next_retry_at", "TIMESTAMP", "", "", "", "", "다음재시도시간"],
        ["settlement_index_queue", "created_at", "TIMESTAMP", "", "", "Y", "NOW()", "생성일시"],
        ["settlement_index_queue", "updated_at", "TIMESTAMP", "", "", "Y", "NOW()", "수정일시"],
        ["settlement_index_queue", "processed_at", "TIMESTAMP", "", "", "", "", "처리일시"],
        # settlement_schedule_config
        ["settlement_schedule_config", "id", "BIGSERIAL", "Y", "", "Y", "", "ID"],
        ["settlement_schedule_config", "config_key", "VARCHAR(100)", "", "", "Y", "", "설정키 (UNIQUE)"],
        ["settlement_schedule_config", "cron_expression", "VARCHAR(100)", "", "", "Y", "", "크론식"],
        ["settlement_schedule_config", "enabled", "BOOLEAN", "", "", "Y", "TRUE", "활성여부"],
        ["settlement_schedule_config", "description", "VARCHAR(500)", "", "", "", "", "설명"],
        ["settlement_schedule_config", "merchant_id", "BIGINT", "", "", "", "", "판매자ID"],
        ["settlement_schedule_config", "created_at", "TIMESTAMP", "", "", "Y", "NOW()", "생성일시"],
        ["settlement_schedule_config", "updated_at", "TIMESTAMP", "", "", "Y", "NOW()", "수정일시"],
        # products
        ["products", "id", "BIGSERIAL", "Y", "", "Y", "", "상품 ID"],
        ["products", "name", "VARCHAR(200)", "", "", "Y", "", "상품명 (UNIQUE)"],
        ["products", "description", "TEXT", "", "", "", "", "상품 설명"],
        ["products", "price", "DECIMAL(10,2)", "", "", "Y", "", "가격 (>=0)"],
        ["products", "stock_quantity", "INTEGER", "", "", "Y", "0", "재고 (>=0)"],
        ["products", "status", "VARCHAR(20)", "", "", "Y", "ACTIVE", "상태"],
        ["products", "category_id", "BIGINT", "", "categories(id)", "", "", "카테고리"],
        ["products", "created_at", "TIMESTAMP", "", "", "Y", "CURRENT_TIMESTAMP", "생성일시"],
        ["products", "updated_at", "TIMESTAMP", "", "", "Y", "CURRENT_TIMESTAMP", "수정일시"],
        # categories
        ["categories", "id", "BIGSERIAL", "Y", "", "Y", "", "카테고리 ID"],
        ["categories", "name", "VARCHAR(100)", "", "", "Y", "", "카테고리명 (UNIQUE)"],
        ["categories", "description", "VARCHAR(500)", "", "", "", "", "설명"],
        ["categories", "parent_id", "BIGINT", "", "categories(id)", "", "", "상위카테고리"],
        ["categories", "display_order", "INT", "", "", "Y", "0", "표시순서"],
        ["categories", "is_active", "BOOLEAN", "", "", "Y", "TRUE", "활성여부"],
        ["categories", "created_at", "TIMESTAMP", "", "", "Y", "CURRENT_TIMESTAMP", "생성일시"],
        ["categories", "updated_at", "TIMESTAMP", "", "", "Y", "CURRENT_TIMESTAMP", "수정일시"],
        # tags
        ["tags", "id", "BIGSERIAL", "Y", "", "Y", "", "태그 ID"],
        ["tags", "name", "VARCHAR(50)", "", "", "Y", "", "태그명 (UNIQUE)"],
        ["tags", "color", "VARCHAR(7)", "", "", "", "#6B7280", "색상"],
        ["tags", "created_at", "TIMESTAMP", "", "", "Y", "CURRENT_TIMESTAMP", "생성일시"],
        # product_tags
        ["product_tags", "product_id", "BIGINT", "Y(복합)", "products(id)", "Y", "", "상품"],
        ["product_tags", "tag_id", "BIGINT", "Y(복합)", "tags(id)", "Y", "", "태그"],
        ["product_tags", "created_at", "TIMESTAMP", "", "", "Y", "CURRENT_TIMESTAMP", "생성일시"],
        # ecommerce_categories
        ["ecommerce_categories", "id", "BIGSERIAL", "Y", "", "Y", "", "카테고리 ID"],
        ["ecommerce_categories", "name", "VARCHAR(200)", "", "", "Y", "", "카테고리명"],
        ["ecommerce_categories", "slug", "VARCHAR(300)", "", "", "Y", "", "슬러그 (UNIQUE)"],
        ["ecommerce_categories", "parent_id", "BIGINT", "", "self(id)", "", "", "상위카테고리"],
        ["ecommerce_categories", "depth", "INT", "", "", "Y", "0", "깊이 (0~2)"],
        ["ecommerce_categories", "sort_order", "INT", "", "", "Y", "0", "정렬순서"],
        ["ecommerce_categories", "is_active", "BOOLEAN", "", "", "Y", "TRUE", "활성여부"],
        ["ecommerce_categories", "created_at", "TIMESTAMP", "", "", "Y", "CURRENT_TIMESTAMP", "생성일시"],
        ["ecommerce_categories", "updated_at", "TIMESTAMP", "", "", "Y", "CURRENT_TIMESTAMP", "수정일시"],
        ["ecommerce_categories", "deleted_at", "TIMESTAMP", "", "", "", "", "soft delete"],
        # product_ecommerce_categories
        ["product_ecommerce_categories", "product_id", "BIGINT", "Y(복합)", "products(id)", "Y", "", "상품"],
        ["product_ecommerce_categories", "category_id", "BIGINT", "Y(복합)", "ecommerce_categories(id)", "Y", "", "카테고리"],
        ["product_ecommerce_categories", "created_at", "TIMESTAMP", "", "", "Y", "CURRENT_TIMESTAMP", "생성일시"],
        # product_images
        ["product_images", "id", "BIGSERIAL", "Y", "", "Y", "", "이미지 ID"],
        ["product_images", "product_id", "BIGINT", "", "products(id)", "Y", "", "상품"],
        ["product_images", "original_file_name", "VARCHAR(255)", "", "", "Y", "", "원본파일명"],
        ["product_images", "stored_file_name", "VARCHAR(255)", "", "", "Y", "", "저장파일명"],
        ["product_images", "file_path", "VARCHAR(500)", "", "", "Y", "", "파일경로"],
        ["product_images", "url", "VARCHAR(500)", "", "", "Y", "", "URL"],
        ["product_images", "content_type", "VARCHAR(100)", "", "", "Y", "", "콘텐츠타입"],
        ["product_images", "size_bytes", "BIGINT", "", "", "Y", "", "파일크기"],
        ["product_images", "width", "INT", "", "", "", "", "너비"],
        ["product_images", "height", "INT", "", "", "", "", "높이"],
        ["product_images", "checksum", "VARCHAR(64)", "", "", "", "", "체크섬"],
        ["product_images", "is_primary", "BOOLEAN", "", "", "Y", "FALSE", "대표이미지"],
        ["product_images", "order_index", "INT", "", "", "Y", "0", "정렬순서"],
        ["product_images", "created_at", "TIMESTAMP", "", "", "Y", "CURRENT_TIMESTAMP", "생성일시"],
        ["product_images", "updated_at", "TIMESTAMP", "", "", "Y", "CURRENT_TIMESTAMP", "수정일시"],
        ["product_images", "deleted_at", "TIMESTAMP", "", "", "", "", "soft delete"],
        # password_reset_tokens
        ["password_reset_tokens", "id", "BIGSERIAL", "Y", "", "Y", "", "ID"],
        ["password_reset_tokens", "user_id", "BIGINT", "", "users(id)", "Y", "", "사용자"],
        ["password_reset_tokens", "token", "VARCHAR(255)", "", "", "Y", "", "토큰 (UNIQUE)"],
        ["password_reset_tokens", "expiry_date", "TIMESTAMP", "", "", "Y", "", "만료일시"],
        ["password_reset_tokens", "used", "BOOLEAN", "", "", "Y", "FALSE", "사용여부"],
        ["password_reset_tokens", "created_at", "TIMESTAMP", "", "", "Y", "CURRENT_TIMESTAMP", "생성일시"],
        # reviews
        ["reviews", "id", "BIGSERIAL", "Y", "", "Y", "", "리뷰 ID"],
        ["reviews", "product_id", "BIGINT", "", "products(id)", "Y", "", "상품"],
        ["reviews", "user_id", "BIGINT", "", "users(id)", "Y", "", "작성자"],
        ["reviews", "rating", "SMALLINT", "", "", "Y", "", "별점 (1~5)"],
        ["reviews", "content", "TEXT", "", "", "", "", "리뷰내용"],
        ["reviews", "created_at", "TIMESTAMP", "", "", "Y", "NOW()", "작성일시"],
        ["reviews", "updated_at", "TIMESTAMP", "", "", "Y", "NOW()", "수정일시"],
        # coupons
        ["coupons", "id", "BIGSERIAL", "Y", "", "Y", "", "쿠폰 ID"],
        ["coupons", "code", "VARCHAR(50)", "", "", "Y", "", "쿠폰코드 (UNIQUE)"],
        ["coupons", "type", "VARCHAR(20)", "", "", "Y", "", "유형 (FIXED/PERCENTAGE)"],
        ["coupons", "discount_value", "NUMERIC(10,2)", "", "", "Y", "", "할인값 (>0)"],
        ["coupons", "min_order_amount", "NUMERIC(10,2)", "", "", "Y", "0", "최소주문금액"],
        ["coupons", "max_uses", "INT", "", "", "Y", "1", "최대사용횟수"],
        ["coupons", "used_count", "INT", "", "", "Y", "0", "사용횟수"],
        ["coupons", "expires_at", "TIMESTAMP", "", "", "", "", "만료일시"],
        ["coupons", "is_active", "BOOLEAN", "", "", "Y", "TRUE", "활성여부"],
        ["coupons", "created_at", "TIMESTAMP", "", "", "Y", "NOW()", "생성일시"],
        ["coupons", "updated_at", "TIMESTAMP", "", "", "Y", "NOW()", "수정일시"],
        # coupon_usages
        ["coupon_usages", "id", "BIGSERIAL", "Y", "", "Y", "", "ID"],
        ["coupon_usages", "coupon_id", "BIGINT", "", "coupons(id)", "Y", "", "쿠폰"],
        ["coupon_usages", "user_id", "BIGINT", "", "users(id)", "Y", "", "사용자"],
        ["coupon_usages", "order_id", "BIGINT", "", "orders(id)", "", "", "주문"],
        ["coupon_usages", "used_at", "TIMESTAMP", "", "", "Y", "NOW()", "사용일시"],
    ]
    write_sheet(ws, headers, data, "테이블 정의서")
    wb.save(os.path.join(DOCS, "설계", "테이블 정의서.xlsx"))
    print(f"  [OK] 테이블 정의서.xlsx - 20개 테이블, {len(data)}개 컬럼")


def update_api_spec():
    wb = Workbook()
    ws = wb.active
    headers = ["Domain", "HTTP Method", "Endpoint", "Method", "Request", "Response", "Auth", "Controller"]
    rows = [
        ["common", "GET", "/", "root()", "-", "Map<String,String>", "N", "RootController"],
        ["user", "POST", "/auth/login", "login()", "LoginRequest", "LoginResponse", "N", "AuthController"],
        ["user", "POST", "/users", "createUser()", "CreateUserRequest", "UserResponse", "N", "UserController"],
        ["user", "GET", "/users/{id}", "getUser()", "PathVar: id", "UserResponse", "N", "UserController"],
        ["user", "GET", "/users/admin/all", "getAllUsers()", "-", "List<UserResponse>", "N", "UserController"],
        ["user", "POST", "/users/password-reset/request", "requestPasswordReset()", "PasswordResetRequestDto", "Map", "N", "UserController"],
        ["user", "POST", "/users/password-reset/confirm", "resetPassword()", "ResetPasswordDto", "Map", "N", "UserController"],
        ["category", "GET", "/admin/categories", "getAllCategories()", "-", "List<EcommerceCategoryResponse>", "ADMIN", "AdminEcommerceCategoryController"],
        ["category", "POST", "/admin/categories", "createCategory()", "EcommerceCategoryRequest", "EcommerceCategoryResponse", "ADMIN", "AdminEcommerceCategoryController"],
        ["category", "PUT", "/admin/categories/{id}", "updateCategory()", "EcommerceCategoryUpdateRequest", "EcommerceCategoryResponse", "ADMIN", "AdminEcommerceCategoryController"],
        ["category", "PATCH", "/admin/categories/{id}/move", "moveCategory()", "CategoryMoveRequest", "EcommerceCategoryResponse", "ADMIN", "AdminEcommerceCategoryController"],
        ["category", "PATCH", "/admin/categories/{id}/sort", "changeSortOrder()", "CategorySortRequest", "EcommerceCategoryResponse", "ADMIN", "AdminEcommerceCategoryController"],
        ["category", "PATCH", "/admin/categories/{id}/activate", "activateCategory()", "-", "EcommerceCategoryResponse", "ADMIN", "AdminEcommerceCategoryController"],
        ["category", "PATCH", "/admin/categories/{id}/deactivate", "deactivateCategory()", "-", "EcommerceCategoryResponse", "ADMIN", "AdminEcommerceCategoryController"],
        ["category", "DELETE", "/admin/categories/{id}", "deleteCategory()", "-", "Void", "ADMIN", "AdminEcommerceCategoryController"],
        ["category", "GET", "/categories", "getActiveCategories()", "-", "List<EcommerceCategoryResponse>", "N", "PublicEcommerceCategoryController"],
        ["category", "GET", "/categories/{slug}", "getCategoryBySlug()", "PathVar: slug", "EcommerceCategoryResponse", "N", "PublicEcommerceCategoryController"],
        ["product", "POST", "/api/categories", "createCategory()", "CreateCategoryRequest", "CategoryResponse", "N", "CategoryController"],
        ["product", "GET", "/api/categories/{id}", "getCategory()", "PathVar: id", "CategoryResponse", "N", "CategoryController"],
        ["product", "GET", "/api/categories", "getAllCategories()", "-", "List<CategoryResponse>", "N", "CategoryController"],
        ["product", "GET", "/api/categories/active", "getActiveCategories()", "-", "List<CategoryResponse>", "N", "CategoryController"],
        ["product", "GET", "/api/categories/root", "getRootCategories()", "-", "List<CategoryResponse>", "N", "CategoryController"],
        ["product", "GET", "/api/categories/parent/{parentId}", "getSubCategories()", "PathVar: parentId", "List<CategoryResponse>", "N", "CategoryController"],
        ["product", "PUT", "/api/categories/{id}", "updateCategory()", "UpdateCategoryRequest", "CategoryResponse", "N", "CategoryController"],
        ["product", "POST", "/api/categories/{id}/activate", "activateCategory()", "-", "Void", "N", "CategoryController"],
        ["product", "POST", "/api/categories/{id}/deactivate", "deactivateCategory()", "-", "Void", "N", "CategoryController"],
        ["product", "POST", "/api/products", "createProduct()", "CreateProductRequest", "ProductResponse", "N", "ProductController"],
        ["product", "GET", "/api/products/{productId}", "getProduct()", "PathVar: productId", "ProductResponse", "N", "ProductController"],
        ["product", "GET", "/api/products", "getAllProducts()", "-", "List<ProductResponse>", "N", "ProductController"],
        ["product", "GET", "/api/products/status/{status}", "getProductsByStatus()", "PathVar: status", "List<ProductResponse>", "N", "ProductController"],
        ["product", "GET", "/api/products/available", "getAvailableProducts()", "-", "List<ProductResponse>", "N", "ProductController"],
        ["product", "PUT", "/api/products/{id}/info", "updateProductInfo()", "UpdateProductInfoRequest", "ProductResponse", "N", "ProductController"],
        ["product", "PUT", "/api/products/{id}/price", "updateProductPrice()", "UpdateProductPriceRequest", "ProductResponse", "N", "ProductController"],
        ["product", "PUT", "/api/products/{id}/stock", "updateProductStock()", "UpdateProductStockRequest", "ProductResponse", "N", "ProductController"],
        ["product", "POST", "/api/products/{id}/activate", "activateProduct()", "-", "ProductResponse", "N", "ProductController"],
        ["product", "POST", "/api/products/{id}/deactivate", "deactivateProduct()", "-", "ProductResponse", "N", "ProductController"],
        ["product", "POST", "/api/products/{id}/discontinue", "discontinueProduct()", "-", "ProductResponse", "N", "ProductController"],
        ["product", "POST", "/admin/products/{id}/images", "uploadImages()", "MultipartFile[]", "List<ProductImageResponse>", "ADMIN", "ProductImageController"],
        ["product", "GET", "/admin/products/{id}/images", "getImages()", "-", "List<ProductImageResponse>", "ADMIN", "ProductImageController"],
        ["product", "PATCH", "/admin/products/{id}/images/{imgId}/primary", "setPrimaryImage()", "-", "ProductImageResponse", "ADMIN", "ProductImageController"],
        ["product", "PATCH", "/admin/products/{id}/images/reorder", "reorderImages()", "ImageReorderRequest", "List<ProductImageResponse>", "ADMIN", "ProductImageController"],
        ["product", "DELETE", "/admin/products/{id}/images/{imgId}", "deleteImage()", "-", "Void", "ADMIN", "ProductImageController"],
        ["product", "POST", "/api/tags", "createTag()", "CreateTagRequest", "TagResponse", "N", "TagController"],
        ["product", "GET", "/api/tags/{id}", "getTag()", "PathVar: id", "TagResponse", "N", "TagController"],
        ["product", "GET", "/api/tags", "getAllTags()", "-", "List<TagResponse>", "N", "TagController"],
        ["product", "GET", "/api/tags/product/{productId}", "getTagsByProduct()", "PathVar: productId", "List<TagResponse>", "N", "TagController"],
        ["product", "PUT", "/api/tags/{id}", "updateTag()", "UpdateTagRequest", "TagResponse", "N", "TagController"],
        ["product", "DELETE", "/api/tags/{id}", "deleteTag()", "-", "Void", "N", "TagController"],
        ["product", "POST", "/api/tags/product/{pId}/tag/{tId}", "addTagToProduct()", "-", "Void", "N", "TagController"],
        ["product", "DELETE", "/api/tags/product/{pId}/tag/{tId}", "removeTagFromProduct()", "-", "Void", "N", "TagController"],
        ["coupon", "POST", "/coupons", "createCoupon()", "CouponCreateRequest", "CouponResponse", "N", "CouponController"],
        ["coupon", "GET", "/coupons", "getAllCoupons()", "-", "List<CouponResponse>", "N", "CouponController"],
        ["coupon", "GET", "/coupons/{code}/validate", "validateCoupon()", "QueryParam: userId,amount", "CouponValidateResponse", "N", "CouponController"],
        ["coupon", "POST", "/coupons/{code}/use", "useCoupon()", "CouponUseRequest", "Void", "N", "CouponController"],
        ["order", "POST", "/orders", "createOrder()", "CreateOrderRequest", "OrderResponse", "N", "OrderController"],
        ["order", "GET", "/orders/{id}", "getOrder()", "PathVar: id", "OrderResponse", "N", "OrderController"],
        ["order", "GET", "/orders/user/{userId}", "getUserOrders()", "PathVar: userId", "List<OrderResponse>", "N", "OrderController"],
        ["order", "GET", "/orders/admin/all", "getAllOrders()", "-", "List<OrderResponse>", "N", "OrderController"],
        ["order", "PATCH", "/orders/{id}/cancel", "cancelOrder()", "PathVar: id", "OrderResponse", "N", "OrderController"],
        ["payment", "POST", "/payments", "createPayment()", "PaymentRequest", "PaymentResponse", "N", "PaymentController"],
        ["payment", "PATCH", "/payments/{id}/authorize", "authorizePayment()", "PathVar: id", "PaymentResponse", "N", "PaymentController"],
        ["payment", "PATCH", "/payments/{id}/capture", "capturePayment()", "PathVar: id", "PaymentResponse", "N", "PaymentController"],
        ["payment", "PATCH", "/payments/{id}/refund", "refundPayment()", "PathVar: id", "PaymentResponse", "N", "PaymentController"],
        ["payment", "GET", "/payments/{id}", "getPayment()", "PathVar: id", "PaymentResponse", "N", "PaymentController"],
        ["payment", "POST", "/payments/toss/confirm", "confirmTossPayment()", "TossPaymentConfirmRequest", "PaymentResponse", "N", "PaymentController"],
        ["payment", "POST", "/payments/toss/cart/confirm", "confirmTossCartPayment()", "TossCartConfirmRequest", "List<PaymentResponse>", "N", "PaymentController"],
        ["review", "POST", "/reviews", "createReview()", "ReviewRequest", "ReviewResponse", "N", "ReviewController"],
        ["review", "GET", "/reviews/product/{productId}", "getProductReviews()", "PathVar: productId", "List<ReviewResponse>", "N", "ReviewController"],
        ["review", "GET", "/reviews/user/{userId}", "getUserReviews()", "PathVar: userId", "List<ReviewResponse>", "N", "ReviewController"],
        ["review", "PUT", "/reviews/{id}", "updateReview()", "ReviewUpdateRequest", "ReviewResponse", "N", "ReviewController"],
        ["review", "DELETE", "/reviews/{id}", "deleteReview()", "QueryParam: userId", "Void", "N", "ReviewController"],
        ["settlement", "GET", "/settlements/{id}", "getSettlement()", "PathVar: id", "SettlementResponse", "N", "SettlementController"],
        ["settlement", "GET", "/settlements/payment/{paymentId}", "getSettlementByPaymentId()", "PathVar: paymentId", "SettlementResponse", "N", "SettlementController"],
        ["settlement", "GET", "/settlements/{id}/pdf", "downloadSettlementPdf()", "PathVar: id", "byte[] (PDF)", "N", "SettlementController"],
        ["settlement", "GET", "/api/settlements/search", "search()", "QueryParams: multiple", "SettlementPageResponse", "N", "SettlementSearchController"],
        ["game", "GET", "/games/baduk", "baduk()", "-", "View: baduk", "N", "GameController"],
        ["game", "GET", "/games/omok", "omok()", "-", "View: omok", "N", "GameController"],
    ]
    write_sheet(ws, headers, rows, "API 명세서")
    wb.save(os.path.join(DOCS, "설계", "API 명세서.xlsx"))
    print(f"  [OK] API 명세서.xlsx - {len(rows)}개 엔드포인트")


if __name__ == "__main__":
    print("=== 산출물 업데이트 ===")
    update_table_definition()
    update_api_spec()
    print("=== 완료 ===")
